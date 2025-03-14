
import time

import chardet
import requests
import re
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from plugins.TideFingercms import Tide_cms
from plugins.ipinfo import IP_info


def title(scan_url_port):
    try:
        # 请求网页
        r = requests.get(scan_url_port, timeout=10)
        # 检测页面编码
        r_detectencode = chardet.detect(r.content)
        actual_encode = r_detectencode['encoding']
        # 使用正则表达式提取<title>标签内容
        response = re.findall(b'<title>(.*?)</title>', r.content, re.S)
        # 如果没有找到<title>，则打印消息
        if not response:
            print('[*] Website: ' + scan_url_port + '\t\t' + '\n')
        else:
            cms_list = Tide_cms(scan_url_port)
            print(scan_url_port+":"+"".join(cms_list))
            if cms_list == []:
                cms_list.append("unknown")
            #print(cms_list)
            cms = "".join(cms_list)

            #IP_info()

            # 如果找到了<title>标签的内容，进行解码
            res = response[0].decode(actual_encode)  # 先解码为字符串
            banner = r.headers.get('server', 'Unknown')  # 防止 'server' 头缺失,web服务器信息
            code = str(r.status_code)
            #print('[*] Website: (url:)' + scan_url_port + '\t\t(banner)' + banner + '\t\t' + 'Title: ' + res +'\t\t' + "code: " + code +'\n')
            file_path = "output.xlsx"
            wb = load_workbook(file_path)
            ws = wb.active
            last_row = ws.max_row
            if last_row==1:
                ws.cell(row=last_row, column=1, value="url")
                ws.cell(row=last_row, column=2, value="banner")
                ws.cell(row=last_row, column=3, value="title")
                ws.cell(row=last_row, column=4, value="code")
                ws.cell(row=last_row, column=4, value="cms")
            ws.cell(row=last_row + 1, column=1, value=scan_url_port)
            ws.cell(row=last_row + 1, column=2, value=banner)
            ws.cell(row=last_row + 1, column=3, value=res)
            ws.cell(row=last_row + 1, column=4, value=code)
            ws.cell(row=last_row + 1, column=4, value="".join(cms))
            wb.save(file_path)
    except Exception as e:
        # 捕获并打印异常
        pass

def start_title(subdomain_list):
    for i in subdomain_list:
        title("http://"+i)
"""
    with open('srcTargrt.txt','r') as file:
        for line in file:
            line = line.strip()
            if line:
                #title("http://"+line)
"""
